"""Parse XLSX files into structural elements."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .elements import Element, HeadingElement, KVPairElement, TableRowElement


def _is_likely_kv_layout(ws, max_sample: int = 20) -> bool:
    """Detect if a sheet uses a key-value layout.

    Matches sheets where column A has labels and column B has values.
    Allows extra columns (C, D, ...) as long as they are mostly empty.
    """
    col_count = ws.max_column
    if not col_count or col_count < 2:
        return False

    row_count = ws.max_row or max_sample
    sample_limit = min(row_count, max_sample)

    keys = []
    extra_col_filled = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=1, max_row=sample_limit, values_only=True):
        vals = list(row)
        total_rows += 1
        k = vals[0]
        if k is not None:
            keys.append(str(k).strip())
        if col_count > 2 and any(v is not None and str(v).strip() for v in vals[2:]):
            extra_col_filled += 1

    if len(keys) < 2:
        return False

    unique_ratio = len(set(keys)) / len(keys)
    if unique_ratio < 0.8:
        return False

    if col_count > 2 and total_rows > 0:
        extra_fill_ratio = extra_col_filled / total_rows
        if extra_fill_ratio > 0.3:
            return False

    return True


def _find_header_row(ws, max_scan: int = 10) -> tuple[int, list[str]]:
    """Find the first non-empty row to use as column headers."""
    row_count = ws.max_row or max_scan
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(row_count, max_scan), values_only=True),
        start=1,
    ):
        values = [str(c).strip() if c is not None else "" for c in row]
        non_empty = sum(1 for v in values if v)
        if non_empty >= 2:
            return row_idx, values
    return 1, []


def parse_xlsx(path: Path) -> list[Element]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    elements: list[Element] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        elements.append(
            HeadingElement(text=sheet_name, level=1, sheet_name=sheet_name)
        )

        if _is_likely_kv_layout(ws):
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                k = str(vals[0]).strip() if vals[0] is not None else ""
                v = str(vals[1]).strip() if len(vals) > 1 and vals[1] is not None else ""
                if not k and not v:
                    continue
                if k and not v:
                    elements.append(
                        HeadingElement(text=k, level=2, sheet_name=sheet_name)
                    )
                elif k:
                    elements.append(
                        KVPairElement(key=k, value=v, sheet_name=sheet_name)
                    )
        else:
            header_row_idx, headers = _find_header_row(ws)

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx <= header_row_idx:
                    continue

                cells = [str(c).strip() if c is not None else "" for c in row]

                if not any(cells):
                    continue

                non_empty_count = sum(1 for c in cells if c)
                is_section_break = non_empty_count == 1 and cells[0] != ""

                elements.append(
                    TableRowElement(
                        cells=cells,
                        headers=headers,
                        is_section_break=is_section_break,
                        sheet_name=sheet_name,
                        row_index=row_idx,
                    )
                )

    wb.close()
    return elements
